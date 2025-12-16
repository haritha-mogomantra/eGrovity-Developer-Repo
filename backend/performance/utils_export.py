# ===========================================================
# performance/utils_export.py
# ===========================================================
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
from django.utils import timezone


# ===========================================================
# Excel Export Utility
# ===========================================================
def generate_excel_report(evaluations, filename="performance_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Report"

    headers = [
        "Emp ID", "Employee Name", "Department", "Manager",
        "Week", "Year", "Total Score", "Average Score (%)", "Rank", "Remarks"
    ]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for e in evaluations:
        ws.append([
            e.employee.user.emp_id,
            f"{e.employee.user.first_name} {e.employee.user.last_name}",
            e.department.name if e.department else "-",
            e.employee.manager.user.first_name if e.employee.manager else "-",
            e.week_number,
            e.year,
            e.total_score,
            round(e.average_score, 2),
            e.rank,
            e.remarks or "",
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ===========================================================
# PDF Export Utility (Single Employee Report)
# ===========================================================
def generate_pdf_report(employee, evaluations):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        title=f"{employee.user.emp_id}_Performance_Report",
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>Employee Performance Report</b>", styles["Title"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"<b>Employee:</b> {employee.user.first_name} {employee.user.last_name}", styles["Normal"]))
    story.append(Paragraph(f"<b>Department:</b> {employee.department.name if employee.department else '-'}", styles["Normal"]))
    story.append(Paragraph(f"<b>Generated on:</b> {timezone.now().strftime('%d %b %Y, %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Table data
    data = [["Week", "Year", "Total Score", "Average (%)", "Rank", "Remarks"]]
    for e in evaluations:
        data.append([
            e.week_number,
            e.year,
            e.total_score,
            f"{round(e.average_score, 2)}%",
            e.rank,
            e.remarks or "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976D2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))

    story.append(table)
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    filename = f"{employee.user.emp_id}_performance_report.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
