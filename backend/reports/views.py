# ===============================================
# reports/views.py
# ===============================================
# Handles:
# - Weekly Consolidated Report
# - Monthly Consolidated Report
# - Manager-Wise Report
# - Department-Wise Report
# - CSV Export
# - Excel Export (Weekly + Monthly)
# - PDF Export (Print Performance Report)
# ===============================================

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions, status
from django.db.models import Avg, F, Q, Window
from django.db.models.functions import Rank
from django.utils import timezone
from django.http import HttpResponse
from datetime import timedelta
from itertools import chain

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from employee.models import Employee
from performance.models import PerformanceEvaluation
from .models import CachedReport
from .serializers import (
    WeeklyReportSerializer,
    MonthlyReportSerializer,
    ManagerReportSerializer,
    CachedReportSerializer,
)
from masters.models import Master
from datetime import date
from reports.utils.pdf_generator import generate_employee_performance_pdf
from notifications.views import create_report_notification 
from django.db.models import Max

def get_latest_completed_week():
    latest = PerformanceEvaluation.objects.aggregate(
        max_year=Max("year"),
    )

    if not latest["max_year"]:
        today = date.today()
        y, w, _ = today.isocalendar()
        return y, w - 1

    max_year = latest["max_year"]

    max_week = (
        PerformanceEvaluation.objects
        .filter(year=max_year)
        .aggregate(max_week=Max("week_number"))["max_week"]
    )

    return max_year, max_week



# ===========================================================
# Latest Week API (Frontend Week Picker Authority)
# ===========================================================
class LatestWeekView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        # 1️⃣ Latest COMPLETED week (default selection)
        latest_year, latest_week = get_latest_completed_week()

        # 2️⃣ Current RUNNING ISO week (max selectable)
        today = date.today()
        current_year, current_week, _ = today.isocalendar()

        return Response({
            # Default auto-selected week
            "latest_year": latest_year,
            "latest_week": latest_week,
            "latest_iso_week": f"{latest_year}-W{str(latest_week).zfill(2)}",

            # Max selectable week
            "current_year": current_year,
            "current_week": current_week,
            "current_iso_week": f"{current_year}-W{str(current_week).zfill(2)}",
        }, status=status.HTTP_200_OK)



# ===========================================================
# NORMALIZER — Convert backend records to frontend-friendly format
# ===========================================================
def normalize_report_rows(records):
    """
    Converts any report record into frontend-friendly format:
    id, name, department, manager, score, rank
    """

    #  Fetch ALL departments once (code → full name)
    department_map = {
        d.code.upper(): d.name
        for d in Master.objects.filter(master_type="DEPARTMENT", is_active=True)
        if d.code
    }

    normalized = []

    for r in records:

        # Pick score safely
        def pick_score(r):
            if r.get("total_score") is not None:
                return r["total_score"]
            if r.get("avg_score") is not None:
                return r["avg_score"]
            if r.get("average_score") is not None:
                return r["average_score"]
            if r.get("best_week_score") is not None:
                return r["best_week_score"]
            return 0

        score_value = pick_score(r)

        # Raw department value (code or name)
        dept_raw = (
            r.get("department")
            or r.get("department_name")
            or "-"
        )

        # 🔥 Convert department code → full name
        dept_key = str(dept_raw).upper() if dept_raw not in ["-", None] else None
        dept_full = department_map.get(dept_key, dept_raw)

        normalized.append({
            "id": r.get("emp_id") or r.get("id") or "-",
            "name": (
                r.get("employee_full_name")
                or r.get("full_name")
                or r.get("name")
                or "-"
            ),
            "department": dept_full,  # <-- FIXED HERE
            "manager": (
                r.get("manager_full_name")
                or r.get("manager")
                or "-"
            ),
            "score": score_value,
            "rank": r.get("rank") or "-",
        })

    return normalized

# ===========================================================
# 1. WEEKLY CONSOLIDATED REPORT
# ===========================================================
class WeeklyReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            latest_year, latest_week = get_latest_completed_week()

            week = int(request.query_params.get("week", latest_week))
            year = int(request.query_params.get("year", latest_year))

            if year > latest_year or (year == latest_year and week > latest_week):
                return Response(
                    {"message": "Future week selection is not allowed."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            save_cache = request.query_params.get("save_cache", "false").lower() == "true"

            # Fetch weekly performance in one query
            qs = (
                PerformanceEvaluation.objects.filter(week_number=week, year=year)
                .select_related("employee__user", "employee__department")
                .annotate(emp_id=F("employee__user__emp_id"))
                .order_by("-total_score")   # Pre-sort for faster ranking
            )

            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for Week {week}, {year}."},
                    status=status.HTTP_200_OK,
                )

            # Collect employee IDs once
            emp_ids = list(qs.values_list("employee_id", flat=True))

            # 🔥 Prepare response
            result = []
            rank_counter = 1

            for p in qs:
                emp = p.employee
                manager = getattr(emp, "manager", None)

                manager_full = (
                    f"{manager.user.first_name} {manager.user.last_name}".strip()
                    if manager else "-"
                )

                result.append({
                    "emp_id": emp.user.emp_id,
                    "employee_full_name": f"{emp.user.first_name} {emp.user.last_name}".strip(),
                    "department": (
                        emp.department.name
                        if emp.department and emp.department.master_type == "DEPARTMENT"
                        else "-"
                    ),
                    "manager_full_name": manager_full,
                    "total_score": float(p.total_score),
                    "average_score": float(p.average_score),
                    "score": float(p.total_score),
                    "week_number": week,
                    "year": year,
                    "rank": rank_counter,
                    "remarks": p.remarks or "",
                })

                rank_counter += 1

            # Save cache (optional)
            if save_cache:
                CachedReport.objects.update_or_create(
                    report_type="weekly",
                    year=year,
                    week_number=week,
                    defaults={
                        "payload": {"records": normalize_report_rows(result)},
                        "generated_by": request.user
                    },
                )

            return Response(
                {
                    "evaluation_period": f"Week {week}, {year}",
                    "total_records": len(result),
                    "records": normalize_report_rows(result),
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# 2. MONTHLY CONSOLIDATED REPORT
# ===========================================================
class MonthlyReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        """Return monthly average performance summary."""
        try:
            month = int(request.query_params.get("month", timezone.now().month))
            year = int(request.query_params.get("year", timezone.now().year))
            save_cache = request.query_params.get("save_cache", "false").lower() == "true"

            qs = PerformanceEvaluation.objects.filter(
                review_date__month=month, year=year
            ).select_related("employee__user", "employee__department")

            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for {month}/{year}."},
                    status=status.HTTP_200_OK,
                )

            data = []
            employees = Employee.objects.filter(
                id__in=qs.values_list("employee_id", flat=True),
                is_active=True
            )

            for emp in employees.select_related("user", "department"):
                emp_qs = qs.filter(employee=emp)
                if not emp_qs.exists():
                    continue

                avg_score = round(emp_qs.aggregate(avg=Avg("average_score"))["avg"], 2)
                best_week_obj = emp_qs.order_by("-average_score").first()

                created_at = getattr(best_week_obj, "created_at", None)

                manager_obj = getattr(emp, "manager", None)

                manager_full_name = (
                    f"{manager_obj.user.first_name} {manager_obj.user.last_name}".strip()
                    if manager_obj else "-"
                )


                data.append({
                    "emp_id": emp.user.emp_id,
                    "employee_full_name": f"{emp.user.first_name} {emp.user.last_name}".strip(),
                    "department": (
                        emp.department.name
                        if emp.department and emp.department.master_type == "DEPARTMENT"
                        else "-"
                    ),
                    "manager_full_name": manager_full_name, 
                    "month": month,
                    "year": year,
                    "avg_score": float(avg_score or 0.0),
                    "score": float(avg_score or 0),
                    "best_week": best_week_obj.week_number,
                    "best_week_score": float(best_week_obj.average_score or 0.0),
                })

            # -----------------------------------------------------------
            # SORT MONTHLY DATA (Fix #5)
            # -----------------------------------------------------------
            data.sort(
                key=lambda x: (
                    -(x.get("avg_score") or 0),
                    x.get("emp_id") or ""
                )
            )

            # Assign rank after sorting
            for idx, row in enumerate(data, start=1):
                row["rank"] = idx

            if save_cache:
                CachedReport.objects.update_or_create(
                    report_type="monthly",
                    year=year,
                    month=month,
                    defaults={
                        "payload": {"records": normalize_report_rows(data)},
                        "generated_by": request.user,
                    },
                )

            # Create notification for monthly report generation
            try:
                message = f"Monthly performance report generated for {month}/{year}."
                create_report_notification(
                    triggered_by=request.user,
                    report_type="Monthly Report",
                    link=f"/reports/monthly/?month={month}&year={year}",
                    message=message,
                    department=None,
                )
            except Exception as e:
                pass

            serialized = MonthlyReportSerializer(data, many=True).data

            return Response(
            {
                "evaluation_period": f"Month {month}, {year}",
                "total_records": len(serialized),
                "records": normalize_report_rows(serialized),
            },
            status=status.HTTP_200_OK,
            )


        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# ===========================================================
# 3. DEPARTMENT-WISE WEEKLY REPORT (Final)
# ===========================================================
class DepartmentReportView(APIView):
    """Returns department-wise weekly performance report."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        department_name = (
            request.query_params.get("department") 
            or request.query_params.get("department_name")
        )
        latest_year, latest_week = get_latest_completed_week()

        week = int(request.query_params.get("week", latest_week))
        year = int(request.query_params.get("year", latest_year))

        if year > latest_year or (year == latest_year and week > latest_week):
                return Response(
                    {"message": "Future week selection is not allowed."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # FIX: allow ALL departments (empty or ALL_DEPT)
        if department_name in [None, "", "ALL_DEPT"]:
            department_name = None

        try:
            # If ALL departments → select every employee
            if department_name is None:
                employees = Employee.objects.filter(is_active=True)
            else:
                employees = Employee.objects.filter(
                    Q(
                        department__master_type="DEPARTMENT",
                        department__name__iexact=department_name
                    ) |
                    Q(
                        department__master_type="DEPARTMENT",
                        department__code__iexact=department_name
                    )
                )

            if not employees.exists():
                return Response({"message": "No employees found."}, status=status.HTTP_200_OK)

            qs = PerformanceEvaluation.objects.filter(
                employee__in=employees, week_number=week, year=year
            ).select_related("employee__user", "employee__department")
            if not qs.exists():
                return Response({"message": f"No performance data found for department {department_name} in Week {week}, {year}."}, status=status.HTTP_200_OK)

            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(),
                    order_by=F("total_score").desc(nulls_last=True)
                )
            )

            records = []
            for perf in ranked:
                emp = perf.employee

                manager_obj = getattr(emp, "manager", None)

                manager_full_name = (
                    emp.manager.user.get_full_name()
                    if emp.manager and emp.manager.user
                    else "-"
                )

                records.append({
                    "department_name": (
                        emp.department.name
                        if emp.department and emp.department.master_type == "DEPARTMENT"
                        else "-"
                    ),
                    "department": (
                        emp.department.name
                        if emp.department and emp.department.master_type == "DEPARTMENT"
                        else "-"
                    ),
                    "emp_id": emp.user.emp_id,
                    "employee_full_name": f"{emp.user.first_name} {emp.user.last_name}".strip(),
                    "manager_full_name": manager_full_name,  
                    "total_score": float(perf.total_score),
                    "average_score": float(perf.average_score),
                    "score": float(perf.total_score),
                    "week_number": week,
                    "year": year,
                    "rank": int(perf.computed_rank),
                    "remarks": perf.remarks or "",
                })


            create_report_notification(
                triggered_by=request.user,
                report_type="Department Weekly Report",
                link=f"/reports/department/?department_name={department_name}&week={week}&year={year}",
                message=f"Department-wise report generated for {department_name} (Week {week}, {year}).",
                department=None,
            )

            return Response(
                {
                    "department_name": department_name,
                    "evaluation_period": f"Week {week}, {year}",
                    "total_employees": len(records),
                    "records": normalize_report_rows(records),
                },
                status=status.HTTP_200_OK,
            )


        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# 4. MANAGER REPORT (Placeholder)
# ===========================================================
class ManagerReportView(APIView):
    """Returns manager-wise weekly performance report."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        manager_id = (
            request.query_params.get("manager")
            or request.query_params.get("manager_id")
        )
        latest_year, latest_week = get_latest_completed_week()

        week = int(request.query_params.get("week", latest_week))
        year = int(request.query_params.get("year", latest_year))

        if year > latest_year or (year == latest_year and week > latest_week):
                return Response(
                    {"message": "Future week selection is not allowed."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        #  FIX: allow ALL managers (empty or ALL_MGR)
        if manager_id in [None, "", "ALL_MGR"]:
            manager_id = None

        try:
            if manager_id is None:
                employees = Employee.objects.filter(is_active=True)
            else:
                manager_qs = Employee.objects.filter(
                    Q(user__emp_id__iexact=manager_id) |
                    Q(user__email__iexact=manager_id) |
                    Q(user__first_name__icontains=manager_id) |
                    Q(user__last_name__icontains=manager_id)
                )

                manager_obj = manager_qs.first()
                if not manager_obj:
                    return Response({"error": "Manager not found."}, status=status.HTTP_404_NOT_FOUND)

                employees = Employee.objects.filter(
                    manager=manager_obj,
                    is_active=True
                )

            if not employees.exists():
                return Response({"message": "No employees found under this manager."}, status=status.HTTP_200_OK)

            qs = PerformanceEvaluation.objects.filter(
                employee__in=employees, week_number=week, year=year
            ).select_related("employee__user", "employee__department")

            if not qs.exists():
                return Response({"message": f"No performance data for Week {week}, {year}."}, status=status.HTTP_200_OK)

            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(),
                    order_by=F("total_score").desc(nulls_last=True)
                )
            )

            records = []
            for perf in ranked:
                emp = perf.employee

                manager_obj = getattr(emp, "manager", None)

                manager_full_name = (
                    f"{manager_obj.user.first_name} {manager_obj.user.last_name}".strip()
                    if manager_obj else "-"
                )

                records.append({
                    "emp_id": emp.user.emp_id,
                    "employee_full_name": f"{emp.user.first_name} {emp.user.last_name}".strip(),
                    "department": (
                        emp.department.name
                        if emp.department and emp.department.master_type == "DEPARTMENT"
                        else "-"
                    ),
                    "manager_full_name": manager_full_name, 
                    "total_score": float(perf.total_score),
                    "average_score": float(perf.average_score),
                    "score": float(perf.total_score),
                    "week_number": week,
                    "year": year,
                    "rank": int(perf.computed_rank),
                    "remarks": perf.remarks or "",
                })


            # Serialize the records to ensure consistent rounding + formatting
            serialized_records = ManagerReportSerializer(records, many=True).data

            return Response({
                "manager_id": manager_id,
                "evaluation_period": f"Week {week}, {year}",
                "total_employees": len(serialized_records),
                "records": normalize_report_rows(serialized_records)
            }, status=status.HTTP_200_OK)


        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# 5. EXCEL EXPORT
# ===========================================================
class ExportWeeklyExcelView(APIView):
    """Exports weekly performance data to Excel."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            latest_year, latest_week = get_latest_completed_week()

            week = int(request.query_params.get("week", latest_week))
            year = int(request.query_params.get("year", latest_year))

            if year > latest_year or (year == latest_year and week > latest_week):
                return Response(
                    {"message": "Future week selection is not allowed."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            qs = (
                PerformanceEvaluation.objects.filter(week_number=week, year=year)
                .select_related("employee__user", "employee__department")
                .annotate(emp_id=F("employee__user__emp_id"))
            )

            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for Week {week}, {year}."},
                    status=status.HTTP_200_OK,
                )

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Week_{week}_{year}"

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = [
                "Emp ID",
                "Employee Name",
                "Department",
                "Total Score",
                "Average Score",
                "Rank",
                "Remarks",
            ]
            ws.append(headers)

            # Apply header styling
            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border


            ranked = qs.annotate(
                computed_rank=Window(
                    expression=Rank(),
                    order_by=F("total_score").desc(nulls_last=True)
                )
            )


            for perf in ranked:
                manager_obj = getattr(perf.employee, "manager", None)
                manager_full_name = (
                    f"{manager_obj.user.first_name} {manager_obj.user.last_name}".strip()
                    if manager_obj else "-"
                )

                ws.append(
                    [
                        perf.employee.user.emp_id,
                        f"{perf.employee.user.first_name} {perf.employee.user.last_name}",
                        (
                            perf.department.name
                            if perf.department and perf.department.master_type == "DEPARTMENT"
                            else "-"
                        ),
                        manager_full_name,
                        float(perf.total_score),
                        float(perf.average_score),
                        int(perf.computed_rank),
                        perf.remarks or "",
                    ]
                )


            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 3

            # Create response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"Weekly_Performance_Report_Week{week}_{year}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportMonthlyExcelView(APIView):
    """Exports monthly performance summary to Excel."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            month = int(request.query_params.get("month", timezone.now().month))
            year = int(request.query_params.get("year", timezone.now().year))

            qs = PerformanceEvaluation.objects.filter(
                review_date__month=month, year=year
            ).select_related("employee__user", "employee__department")

            if not qs.exists():
                return Response(
                    {"message": f"No performance data found for {month}/{year}."},
                    status=status.HTTP_200_OK,
                )

            wb = Workbook()
            ws = wb.active
            ws.title = f"Month_{month}_{year}"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            headers = [
                "Emp ID",
                "Employee Name",
                "Department",
                "Manager",
                "Average Score",
                "Best Week",
                "Best Week Score",
            ]
            ws.append(headers)

            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

            employees = Employee.objects.filter(
                id__in=qs.values_list("employee_id", flat=True),
                is_active=True
            ).select_related("user", "department")

            # Pre-cache evaluations per employee to avoid repeated filtering
            emp_evaluations = {emp.id: qs.filter(employee=emp) for emp in employees}

            for emp in employees:
                emp_qs = emp_evaluations.get(emp.id)

                if not emp_qs or not emp_qs.exists():
                    continue


                avg_score = round(emp_qs.aggregate(avg=Avg("average_score"))["avg"], 2)
                best_week_obj = emp_qs.order_by("-average_score").first()
                created_at = getattr(best_week_obj, "created_at", None)

                manager_obj = getattr(emp, "manager", None)
                manager_full_name = (
                    f"{manager_obj.user.first_name} {manager_obj.user.last_name}".strip()
                    if manager_obj else "-"
                )

                ws.append(
                    [
                        emp.user.emp_id,
                        f"{emp.user.first_name} {emp.user.last_name}",
                        (
                            emp.department.name
                            if emp.department and emp.department.master_type == "DEPARTMENT"
                            else "-"
                        ),
                        manager_full_name,
                        float(avg_score or 0.0),
                        best_week_obj.week_number,
                        float(best_week_obj.average_score or 0.0),
                    ]
                )



            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 3

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"Monthly_Performance_Report_{month}_{year}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



# ===========================================================
# 6. PRINT PERFORMANCE REPORT (PDF Export)
# ===========================================================
class PrintPerformanceReportView(APIView):
    """
    Generates and returns a downloadable PDF report for an individual employee’s
    weekly performance. Integrates with the ReportLab-based PDF generator utility.
    
    Example:
      GET /api/reports/print/EMP0001/?week=44&year=2025
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, emp_id):
        try:
            latest_year, latest_week = get_latest_completed_week()

            week = int(request.query_params.get("week", latest_week))
            year = int(request.query_params.get("year", latest_year))

            if year > latest_year or (year == latest_year and week > latest_week):
                return Response(
                    {"message": "Future week selection is not allowed."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Fetch employee
            try:
                employee = Employee.objects.select_related("user", "department").get(
                    user__emp_id__iexact=emp_id,
                    is_active=True
                )
            except Employee.DoesNotExist:
                return Response({"error": f"Employee with ID {emp_id} not found."}, status=status.HTTP_404_NOT_FOUND)

            # Fetch performance evaluations for the selected week
            evaluations = PerformanceEvaluation.objects.filter(
                employee=employee, week_number=week, year=year
            ).select_related("employee__user", "employee__department")

            if not evaluations.exists():
                return Response(
                    {"message": f"No performance records found for {emp_id} in Week {week}, {year}."},
                    status=status.HTTP_200_OK,
                )

            # Generate the PDF using the utility
            pdf_response = generate_employee_performance_pdf(employee, evaluations, week=f"Week {week}, {year}")

            # Create a notification for this PDF export
            try:
                create_report_notification(
                    triggered_by=request.user,
                    report_type="Employee Performance PDF",
                    link=request.get_full_path(),
                    message=f"PDF performance report generated for {employee.user.emp_id} ({employee.user.first_name} {employee.user.last_name}).",
                    department=employee.department,
                )
            except Exception as e:
                pass

            return pdf_response

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ===========================================================
# 7. CACHED REPORT MANAGEMENT (List, Archive, Restore)
# ===========================================================
from rest_framework.generics import ListAPIView
from django.shortcuts import get_object_or_404

class CachedReportListView(ListAPIView):
    """Displays list of cached reports (Admin/Manager dashboard)."""
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = CachedReportSerializer

    def get_queryset(self):
        queryset = CachedReport.objects.all().order_by("-generated_at")
        report_type = self.request.query_params.get("report_type")
        if report_type:
            queryset = queryset.filter(report_type__iexact=report_type)
        return queryset


class CachedReportArchiveView(APIView):
    """Archives a cached report (soft delete)."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        report = get_object_or_404(CachedReport, pk=pk)
        report.is_active = False
        report.save(update_fields=["is_active"])
        return Response(
            {"message": f"Report {report.id} archived successfully."},
            status=status.HTTP_200_OK,
        )


class CachedReportRestoreView(APIView):
    """Restores an archived cached report."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        report = get_object_or_404(CachedReport, pk=pk)
        report.is_active = True
        report.save(update_fields=["is_active"])
        return Response(
            {"message": f"Report {report.id} restored successfully."},
            status=status.HTTP_200_OK,
        )
